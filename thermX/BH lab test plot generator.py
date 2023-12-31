# -*- coding: utf-8 -*-
"""
Created on Fri Aug 31 12:34:34 2018

@author: wwei
"""
"""


Open a file dialog window in tkinter using the filedialog method.
Tkinter has a prebuilt dialog window to access files. 
This example is designed to show how you might use a file dialog askopenfilename
and use it in a program.
"""

import osdis as osd


import logging
'''
cwd = os.getcwd()
logging.debug(cwd)
clear = lambda: os.system('cls')
clear()
'''


osd.getOsInfo()
osd.getPID()
osd.getPythonExe()
print(f'The current directory is {osd.getCurDir()}.')
print(f'{osd.getLog()}')
'''
from matplotlib import rc
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter,
                               AutoMinorLocator)
minorLocator = AutoMinorLocator()
import math

from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl import Workbook

from openpyxl.chart import (LineChart,
                            ScatterChart,
                            Reference, 
                            Series)
'''
try: 
    from openpyxl.cell import get_column_letter, column_index_from_string, coordinate_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
'''


Font = "Times New Roman"
Font_size = "11"
#The height needed for the table for single weldment of parameter entry
height = 4
    
Label_height = "1"
Label_width = "19"
Input_width = str( int(Label_width) + 3 )
Label_bd = "2"
Label_relief = "groove"
        
row_ini = 0
column_ini = 0
row_ini_table = 2
ncsm = 4 # num_col_shift_menu

logging.basicConfig(level = logging.DEBUG, 
                    format = ' %(asctime)s -  %(levelname)s -  %(message)s')

logging.debug('Start of program')

#===============================================
def plot_2D(fcase, x, y, xlbl, ylbl):
    
    
#    x = np.linspace(x_min,x_max,points)
#    T = T_inf + Q/P * ( 1 - np.exp(-P*x) ) + ( T_i - T_inf ) * np.exp(-P*x)
    logging.debug(fcase)
    logging.debug(x)
    logging.debug(float(max(y)))
    plt.figure(fcase)
    
    plt.xticks( range(len(x)), x )
    
    plt.yticks(np.arange(0, 250, step=20))
    plt.ylim(0, 250)
#    ax.yaxis.set_minor_locator(minorLocator)
#    plt.scatter(range(len(y)), y )
    plt.plot(y,'ro')
    plt.xlabel(xlbl, fontsize=16)
    plt.ylabel(ylbl, fontsize=16)
    plt.xticks(color='k', size=16)
    plt.yticks(color='k', size=16)
#    plt.xlim(0, np.amax(x) )
#    
    plt.tight_layout()
    plt.show()
    
def GenMatlabPlot():
    ULind = str( Input_UL.get() )  #get the upper left index for plot
    LRind = str( Input_LR.get() )  #get the lower right index for plot
    ULxy = coordinate_from_string( ULind )
    LRxy = coordinate_from_string( LRind )
    
#    logging.debug( ws )
#    logging.debug( ws[ULind] )
#    logging.debug( ws[ULind].value )    
#    logging.debug('End of program')
    
    row_min = ULxy[1]
    row_max = LRxy[1]
    logging.debug( row_min )
    logging.debug( row_max )
    col_min = column_index_from_string( ULxy[0] )
    col_max = column_index_from_string( LRxy[0] )
    logging.debug( col_min )
    logging.debug( col_max )
    rows = {}
    
    
    total_col = ( col_max - col_min ) + 1 
    total_row = ( row_max - row_min ) + 1
    try:
        for j in range(col_min-1, col_max + 1):
            x = []
            y = []
            for i in range(row_min-1, row_max + 1):
                index = str(get_column_letter(j)) + str(i)
    #            logging.debug(index)
                logging.debug(ws[index].value)
                rows[i, j] = ws[index].value
                logging.debug( rows[i, j] )
                xtemp = str(rows[i, col_min-1])
                x.append(xtemp)
                if j != col_min-1:
                    if i == row_min-1:
                        ytemp = str(rows[i,j])
                        y.append(ytemp)
                    elif i > row_min-1:
                        ytemp = float(rows[i,j])
                        y.append(ytemp)
                    
            if j != col_min-1:     
                logging.debug(x[1:total_row+1])
                logging.debug(y[1:total_row+1])
                fcase = y[0]
                xlbl = 'T/C'
                ylbl = 'Temperature (°C)'
                plot_2D(fcase, x[1:total_row+1], y[1:total_row+1], xlbl, ylbl)  

    except KeyError:
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()

#===============================================
def GenPlot():
    ULind = str( Input_UL.get() )  #get the upper left index for plot
    LRind = str( Input_LR.get() )  #get the lower right index for plot
    ULxy = coordinate_from_string( ULind )
    LRxy = coordinate_from_string( LRind )
    
    logging.debug( ws )
#    logging.debug( wb )
#    logging.debug( ws[ULind].value )    
#    logging.debug('End of program')
    
    row_min = ULxy[1]
    row_max = LRxy[1]
    logging.debug( row_min )
    logging.debug( row_max )
    col_min = column_index_from_string( ULxy[0] )
    col_max = column_index_from_string( LRxy[0] )
    logging.debug( col_min )
    logging.debug( col_max )
    rows = {}
    
    total_cols = ( col_max - col_min ) + 1 
    try:
        for j in range(col_min-1, col_max + 1):        
            for i in range(row_min-1, row_max + 1):
                index = str(get_column_letter(j)) + str(i)
    #                logging.debug(index)
                logging.debug(ws[index].value)
                rows[i, j] = ws[index].value
                logging.debug( rows[i, j] )
            
        
    
        wbb = Workbook() # Create a new XlsxWriter Workbook object.
        wss = wbb.active  # This gives you the currently active worksheet
           
        for j in range(col_min-1, col_max + 1):       
            for i in range(row_min-1, row_max + 1):
                index = str(get_column_letter(j)) + str(i)
    #            logging.debug(ws[index].value)
                wss[index].value = rows[i, j]
    #            logging.debug( rows[i, j] )
                
        #Define the data range including case name (at row_min-1)
        data = Reference(wss, min_col=col_min, min_row=row_min-1, 
                              max_col=col_min+1, max_row=row_max)
        logging.debug(data)
    
        chart = LineChart()
        
        chart.title = "Temperature distribution"
        chart.style = 13
        chart.x_axis.title = 'T/C'
        chart.y_axis.title = 'Temperature (°C)'
        
        chart.add_data(data, titles_from_data = True)
        TC = Reference(wss, min_col=col_min-1, min_row=row_min, max_row=row_max)
        #Setting T/C numbers as the x-axis coordinates
        chart.set_categories(TC)
        
        # Style the lines
        for k in range(0, total_cols):
            s = chart.series[k]
            s.marker.symbol = "circle"
#            s.marker.graphicalProperties.line.solidFill = "FF0000"
#            s.graphicalProperties.line.solidFill = "00AAAA"
            s.graphicalProperties.line.noFill = True
            logging.debug(k)
        
    
        wss.add_chart(chart, "A1")
        file_name = 'Temperature distribution_' + Input_PN.get() + '_' + Input_TD.get() + '.xlsx'
        logging.debug(file_name)
        wbb.save(file_name)
        
        logging.debug('Generating the plot successful')
        
        top = Toplevel()
        w = 600
        h = 100 
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Done')

        msg = Message(top, text='Generating the plot in ' + file_name + ' successful!', width=600)
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
        
        
    except KeyError:
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
        
    
    

#===============================================    

def OpenExcel(path):
    
    global ws
#    logging.debug(ws)
    logging.debug(root.filename)
    #openpyxl.load_workbook() function takes in the filename and returns value of the workbook data type.
    wb = openpyxl.load_workbook(root.filename)  
#    logging.debug( wb )
    logging.debug( wb.sheetnames )
#    logging.debug( root.filename )
#    logging.debug( wb.get_active_sheet() )
#    ws_all = wb.active
#    logging.debug( ws_all )
    
    try:
        ws = wb['Results']
    except KeyError:
        
        try:
            ws = wb['Result']
        except KeyError:
            
            top = Toplevel()
            w = 150
            h = 100   
            top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
            top.title('Warning')
    
            msg = Message(top, text="Please confirm if 'Result' or 'Results' in the test file." )
            msg.pack()
    
            button = Button(top, text="Exit", command=top.destroy)
            button.pack()    
   
#===============================================
#This is where we lauch the file manager bar.
def OpenFile():
    
    global path
#    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All Excel Files", "*.xl*;*.xlsx*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )                 

    path = root.filename
    logging.debug(root.filename)
    logging.debug(Input_Path.get())
    logging.debug(ws)
    Input_Path.insert(50, root.filename)
    
    if len(Input_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
    elif len(Input_Path.get()) > 1:
        OpenExcel(path)
        
#    logging.debug(ws)

#===============================================
def img_label(imgPath, r, col):
#    print(imgPath)
    photo = PhotoImage(file = imgPath)
    label = Label(root, image=photo, bg = 'lightgrey')
    label.image = photo
    label.grid(row = r, column = col, columnspan=5, rowspan=2,sticky=W+N+S, padx=5, pady=5)
    
#===============================================
def clear_all():
#    gui.grid_forget()

    for widget in root.winfo_children():      # get all children of root
        if widget.winfo_class() == 'Entry':   # if the class is Entry
            widget.delete(0,END)              # reset its value
        elif widget.winfo_class() == 'Label': # get Label widgets
            widget.config(bg="light grey")         # and change them as well
#===============================================            
def des():
    logging.debug('End of program')
    root.destroy()
            
#===============================================            
if __name__ == "__main__":
    global wids, heis, rx, ry, w, h
    ws = {}
    
    root = Tk() #create one root widget for this program
    # get screen width and height
    wids = root.winfo_screenwidth() # width of the screen
    heis = root.winfo_screenheight() # height of the screen
    rx = root.winfo_x()
    ry = root.winfo_y()
    logging.debug( wids )
    logging.debug( heis )

    
    Title = root.title( "Generating plots of T/C reading (Technical support: ext. 1235)")
    root.configure(background="Dark grey")
    root.geometry("1000x384")
    
    #This button serves as a label to select a file
#    Label(root, text = 'File', font=(Font, Font_size), 
#          height = Label_height, width = Label_width, bd=Label_bd, 
#          relief="groove").grid(row = 0, column = 0, sticky = W)
    Input_Path = Entry(root, width = '100', font=(Font, Font_size))
    Input_Path.grid(row = 0, column = 1, sticky=NSEW)
    Input_Path.insert(50, ' ')    
    
    tk.Button(root, text='Select a test result', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFile
              ).grid(row = 0 , column = 0, sticky = NSEW)
    
    #Enter part number of jacket
    Label(root, text = 'Part number of jacket', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 1, column = 0, sticky = NSEW)
    
    Input_PN = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_PN.grid(row = 1, column = 1, sticky=W)
    Input_PN.insert(10, "UAPMXX0XXXSN-XXX")    
        
    #Enter test date
    Label(root, text = 'Test date (MMDDYYYY)', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 2, column = 0, sticky = NSEW)
    
    Input_TD = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_TD.grid(row = 2, column = 1, sticky=W)
    Input_TD.insert(10, "12312018")   
    
    
    #Enter the index of upper left cell
    Label(root, text = 'Index of upper left cell', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 3, column = 0, sticky = NSEW)
    
    Input_UL = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_UL.grid(row = 3, column = 1, sticky=W)
    Input_UL.insert(10, 'G31')
    
    #Enter the index of lower right cell
    Label(root, text = 'Index of lower right cell', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 4, column = 0, sticky=NSEW)
    
    Input_LR = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_LR.grid(row = 4, column = 1, sticky=W)
    Input_LR.insert(10, 'H35')    
    
    # generate the button for creaing input table (Now)
    Button(root, text = 'Generating plots in an excel file', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=GenPlot).grid(row = 5, column= 0, sticky=NSEW)
    
    Button(root, text = 'Generating MATLAB plots', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=GenMatlabPlot).grid(row = 6, column= 0, sticky=NSEW)

    Buttn_refresh = tk.Button(root, text='Clear all entries', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "Black", bg = 
                       "Red", command=clear_all).grid(row = 7, column = 0, sticky=NSEW)
    
    Buttn_end = tk.Button(root, text='End', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "black", bg = 
                       "red", command=des).grid(row = 8, column = 0, sticky=NSEW)
    
    imgPath = "BH_logo.gif"
    img_label(imgPath, 9, 0)   


    root.mainloop()
    logging.debug('End of program')
    '''