import openpyxl
from openpyxl import Workbook

from openpyxl.chart import (LineChart,
                            ScatterChart,
                            Reference, 
                            Series)
try: 
    from openpyxl.cell import get_column_letter, column_index_from_string#, coordinate_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string#, coordinate_from_string


class RawDataClass():
    def writeToFile(self):
        print('wrte to file')
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        wb.save("sample.xlsx")

    def readFromFile(self, pathFileName):
        print("Read from file at ", pathFileName)

def main():
    rdObj = RawDataClass()
    rdObj.writeToFile()
    #rdObj.readFromFile("C:\Users\williamw\Documents\William\ideas\ScriptIdea\\thermX\\raw_data_file_development")

if __name__ == "__main__":
    main()