from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename

import logging

Font = "Times New Roman"
Font_size = "11"
#The height needed for the table for single weldment of parameter entry
height = 4
    
Label_height = "1"
Label_width = "19"
Input_width = str( int(Label_width) + 3 )
Label_bd = "2"
Label_relief = "groove"
        
row_ini = 0
column_ini = 0
row_ini_table = 2
ncsm = 4 # num_col_shift_menu

import OsdisClass as osdc
osdcObj = osdc.OSDIS()
cwd = osdcObj.getCurDir()

import openpyxl

import RawDataClass as rdc
rdcObj = rdc.RawDataClass()

import TcPlotClass as tpc
tpcObj = tpc.TcPlotClass()

def GenMatlabPlot():
    
    setTCTransit(ws)
    tcLocDic = getTCLoc()
    if (int(Input_mvAvg.get()) >= 5):
        tcAvgList = getMovingAverage(getTCTransit(), 
                                     int(Input_mvAvg.get()))
    tpcObj.temp2Dmap(tcLocDic, tcAvgList)

def OpenFile():
    global path
#    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All Excel Files", "*.xl*;*.xlsx*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )                 

    path = root.filename
    #logging.debug(root.filename)
    #logging.debug(Input_Path.get())
    #logging.debug(ws)
    Input_Path.insert(50, root.filename)
    
    if len(Input_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
    elif len(Input_Path.get()) > 1:
        OpenExcel(path)



def getRowTatal(sheetRD):
    nR = 0
    for row in sheetRD:
        nR+=1
    return nR

def showRawData(sheetRD):
    for row in sheetRD:
        for col in row:
            print(col.value)

def setTCLoc(sheetRD):
    global gDicTCLoc
    gDicTCLoc = dict()
    for row in sheetRD:
        #for col in row:
        coordTmp = []
        coordTmp.append(row[1].value)
        coordTmp.append(row[2].value)
        gDicTCLoc[row[0].value] = coordTmp
 
def getTCLoc():
    return gDicTCLoc

def setTCTransit(sheetRD):
    rowIndMax = getRowTatal(sheetRD)
    global gDicTCTransit
    gDicTCTransit = dict()
    count = 0
    for row in sheetRD:
        if count == 0:
            gDicTCTransit[row[0].value] = row[1].value
        if count >= 34:
            tTransit = []
            for j in range(2,27):
                tTransit.append(row[j].value)            
            gDicTCTransit[row[1].value] = tTransit
        count+=1

def getTCTransit():
    return gDicTCTransit

def getMovingAverage(dicTCtransit, n):
    print("Here getMovingAverage", n)
    print('size ', len(dicTCtransit))
    
    r = 0
    sTmp = [0] * 25
    for key, value in dicTCtransit.items():
        if r in range(len(dicTCtransit)-n, len(dicTCtransit)):
            for j in range(0, len(value)):
                sTmp[j]+=value[j]
        r+=1
    
    tcAvg = []
    for j in range(0, len(value)):
        tcAvg.append(sTmp[j]/n)
    #print(tcAvg)
    return tcAvg
    

def GenPlot():
    pass

def OpenExcel(path):
    
    global ws, wsTC
#    logging.debug(ws)
    logging.debug(root.filename)
    #openpyxl.load_workbook() function takes in the filename and returns value of the workbook data type.
    wb = openpyxl.load_workbook(root.filename)  

#    logging.debug( wb )
    # show the number of sheets/tabs in a single file
    logging.debug(len(wb.sheetnames))
    logging.debug( wb.sheetnames )
#    logging.debug( root.filename )
#    logging.debug( wb.get_active_sheet() )
#    ws_all = wb.active
#    logging.debug( ws_all )
    
    try:
        ws = wb['X05469 SST 2nd heater426,428,42']
        wsTC = wb['TC_number_location']
    except KeyError:
        
        try: #if no file name is found
            ws = wb.worksheets[0]   #raw data
            wsTC = wb.worksheets[1] #TC number location
        except KeyError:
            
            top = Toplevel()
            w = 150
            h = 100   
            top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
            top.title('Warning')
    
            msg = Message(top, text="Please confirm if 'Result' or 'Results' in the test file." )
            msg.pack()
    
            button = Button(top, text="Exit", command=top.destroy)
            button.pack()    

    # show stored raw data
    #showRawData(ws)
    #showRawData(wsTC)
    
    setTCLoc(wsTC) 
    #print(getTCLoc())   

    setTCTransit(ws)
    #print(getTCTransit())

    #if (int(Input_mvAvg.get()) >= 5):
    #    tcAvgList = getMovingAverage(getTCTransit(), int(Input_mvAvg.get()))


#===============================================
def clear_all():
    for widget in root.winfo_children():            # get all children of root
        if widget.winfo_class() == 'Entry':         # if the class is Entry
            widget.delete(0,END)                    # reset its value
        elif widget.winfo_class() == 'Label':       # get Label widgets
            widget.config(bg="light grey")          # and change them as well

#===============================================            
def des():
    logging.debug('End of program')
    root.destroy()

if __name__ == "__main__":
    root = Tk() #create one root widget for this program
    logging.basicConfig(level = logging.DEBUG, 
                    format = ' %(asctime)s -  %(levelname)s -  %(message)s')
    
    global wids, heis, rx, ry, w, h
    ws = {}

    #root = Tk() #create one root widget for this program
    # get screen width and height
    wids = root.winfo_screenwidth() # width of the screen
    heis = root.winfo_screenheight() # height of the screen
    rx = root.winfo_x()
    ry = root.winfo_y()

    #logging.debug( wids )
    #logging.debug( rx )

    Title = root.title( "Automation of generating plots of test T/C reading (Technical support: ext. ????)")
    root.configure(background="Dark grey")
    defSize = str(int(0.5 * wids)) + \
              "x" + \
              str(int(0.5 * heis))
      
    root.geometry(defSize)

    #logging.debug( defSize )

    Input_Path = Entry(root, width = '100', font=(Font, Font_size))
    Input_Path.grid(row = 0, column = 1, sticky=NSEW)
    Input_Path.insert(50, ' ')   

    tk.Button(root, text='Select a test result file', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFile
              ).grid(row = 0 , column = 0, sticky = NSEW)

    #Enter TC config
    '''
    Input_Path2 = Entry(root, width = '100', font=(Font, Font_size))
    Input_Path2.grid(row = 1, column = 1, sticky=NSEW)
    Input_Path2.insert(50, ' ') 

    tk.Button(root, text='Select a TC config file', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFile
              ).grid(row = 1 , column = 0, sticky = NSEW)
    '''
    '''
    Label(root, text = 'TC location ', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 1, column = 0, sticky = NSEW)
    

    Input_PN = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_PN.grid(row = 1, column = 1, sticky=W)
    Input_PN.insert(10, "25")    
    '''

    #Enter average points
    Label(root, text = 'Moving average points in raw data', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 2, column = 0, sticky = NSEW)
    
    Input_mvAvg = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_mvAvg.grid(row = 2, column = 1, sticky=W)
    Input_mvAvg.insert(10, "5")
    
    
    #Enter the index of upper left cell
    Label(root, text = 'Info to be added', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 3, column = 0, sticky = NSEW)
    
    Input_UL = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_UL.grid(row = 3, column = 1, sticky=W)
    Input_UL.insert(10, 'G31')
    '''
    #Enter the index of lower right cell
    Label(root, text = 'Index of lower right cell', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 4, column = 0, sticky=NSEW)
    
    Input_LR = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_LR.grid(row = 4, column = 1, sticky=W)
    Input_LR.insert(10, 'H35')    
    '''
    # generate the button for creaing input table (Now)
    Button(root, text = 'Generating transient plots in an excel file', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=GenPlot).grid(row = 5, column= 0, sticky=NSEW)
    
    Button(root, text = 'Generating 2D temperature plots', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=GenMatlabPlot).grid(row = 6, column= 0, sticky=NSEW)

    Buttn_refresh = tk.Button(root, text='Clear all entries', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "Black", bg = 
                       "yellow", command=clear_all).grid(row = 7, column = 0, sticky=NSEW)
    
    Buttn_end = tk.Button(root, text='End', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "black", bg = 
                       "yellow", command=des).grid(row = 8, column = 0, sticky=NSEW)
    

    root.mainloop()
    logging.debug('End of program')
