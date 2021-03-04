#! python3
"""
• On Windows, the shebang line is #! python3.
• On OS X, the shebang line is #! /usr/bin/env python3.
• On Linux, the shebang line is #! /usr/bin/python3.
"""
import os, logging
cwd1 = os.getcwd()
cwd = cwd1
logging.debug(cwd)
clear = lambda: os.system('cls')
clear()

from matplotlib import rc
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter,
                               AutoMinorLocator)
minorLocator = AutoMinorLocator()
import math
import csv
import pandas as pd
from copy import deepcopy
#import PyPDF2

from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl import Workbook

from openpyxl.chart import (LineChart,
                            ScatterChart,
                            Reference, 
                            Series)

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from openpyxl.styles.borders import Border, Side

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

try: 
    from openpyxl.cell import get_column_letter, column_index_from_string, coordinate_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
    
    
import datetime
import winsound


curDate = datetime.datetime.today().strftime('%m%d%Y')



Font = "Times New Roman"
Font_size = "11"
#The height needed for the table for single weldment of parameter entry
height = 4
    
Label_height = "1"
Label_width = "25"
Input_width = str( int(Label_width) + 3 )
Label_bd = "2"
Label_relief = "groove"
        
row_ini = 0
column_ini = 0
row_ini_table = 2
ncsm = 4 # num_col_shift_menu
tablist = ['Result', 'UTR', 'Setup', 'Chart', 'Duty', 'Data log', 'Picture']

ReTabList = ['', 'Thermocouple (°C)', 'Inlet', 'T/C', 'Outlet', 'Control', 
             'Duty cycle (%)', 'Maximum wattage (W)', 'Power (W)', 'Notes']

SPlist = ['Date of testing', 'Run time (h:mm:ss)', 'Jacket part number', 'Jacket customer part number',
          'Jacket serial number', 'Jacket resistance (Ω)', 'Power input (VAC)', 'Thermocouple', 
          'Temperature controller', 'Data acquisition hardware', 'Data acquisition software', 
          'Set point temperatures (°C)']

# Default Color Index as per 18.8.27 of ECMA Part 4
COLOR_INDEX = (
    '000000', '333399', 'FF0000', '00FF00', '0000FF', #0-4
    'FFFF00', 'FF00FF', '00FFFF', '800000', '008000', #5-9
    '000080', '808000', '800080', '008080', 'C0C0C0', #10-14
    '808080', '9999FF', '993366', 'FFFFCC', 'CCFFFF', #15-19
    '660066', 'FF8080', '0066CC', 'CCCCFF', '000080', #20-24
    'FF00FF', 'FFFF00', '00FFFF', '800080', '800000', #25-29
    '008080', '0000FF', '00CCFF', 'CCFFFF', 'CCFFCC', #30-34
    'FFFF99', '99CCFF', 'FF99CC', 'CC99FF', 'FFCC99', #35-39
    '3366FF', '33CCCC', '99CC00', 'FFCC00', 'FF9900', #40-44
    'FF6600', '666699', '969696', '003366', '339966', #45-49
    '003300', '333300', '993300', '993366', '333333' #50-54
    )                                         #51

## Will remove these definitions in a future release
#BLACK = COLOR_INDEX[0]
#WHITE = COLOR_INDEX[1]
#RED = COLOR_INDEX[2]
#GREEN = COLOR_INDEX[3]
#BLUE = COLOR_INDEX[4]
#YELLOW = COLOR_INDEX[5]
#Magenta = COLOR_INDEX[6]
#Cyan = COLOR_INDEX[7]

global row_min, row_max, col_min, col_max
#initial guess of the row and col numbers of the raw T/C data
row_min = 1
row_max = 100
    
col_min = 1
col_max = 100


logging.basicConfig(level = logging.DEBUG, 
                    format = ' %(asctime)s -  %(levelname)s -  %(message)s')

logging.debug('Start of program')
#===============================================
def readCase():
    global excelCaName, cus_jack_PN 
    
    excelCaName = []
    for row in rows:
        for col in row:
            excelCaName.append(col.get())
            
    for i in range(len(excelCaName)):
        print(i, excelCaName[i])
        
    #This is used to add the jacket customer PN into figures
    cus_jack_PN = JCPN.get()
    print(cus_jack_PN)
    
#===============================================
def popCase():
    
    global rows, caseInp, top, numSP, numC, JCPN, row_max
        # Add a grid
    top=Toplevel(root)
    Label(top, text="Case #", font=(Font, Font_size), 
          height = Label_height, bd=Label_bd, 
          relief="groove").grid(row = row_ini, column = 0, sticky = NSEW)
    
    Label(top, text="     Entry case name (i.e., no flow, 10 SLM, etc.)     ", font=(Font, Font_size), 
          height = Label_height, bd=Label_bd, 
          relief="groove").grid(row = row_ini, column = 1, sticky = NSEW)
    
    print(Input_CT.get())
    numC = int(Input_CT.get())

    #===============================================
    #Creating case entris for the result tab
    cols = []
    rows = []
    for i in range(row_ini+1, numC+1):
        Label(top, text="Case" + str(i), font=(Font, Font_size), 
          height = Label_height, bd=Label_bd, 
          relief="groove").grid(row = row_ini+i, column = 0, sticky = NSEW)
        
#        caseInpt = NameInpt.append("Input_" + str(i))
        caseInp = Entry(top, width = Input_width, font=(Font, Font_size))
        caseInp.grid(row = row_ini+i, column = 1, sticky=NSEW)
        caseInp.insert(10, str(i) + " SLM")
        cols.append(caseInp)
        
    rows.append(cols)
        
#    ===============================================
    Label(top, text="Setup", font=(Font, Font_size), 
          height = Label_height, bd=Label_bd,
          relief="groove").grid(row = numC+1, column = 0, sticky = NSEW, columnspan=2)

    #Creating entris for the setup tab
#    cols = []
#    rows = []
    numSP = 12
    iniSP = numC+2
    
    for i in range(iniSP, iniSP+numSP):
        Label(top, text=SPlist[i - (iniSP)], font=(Font, Font_size), 
          height = Label_height, bd=Label_bd, 
          relief="groove").grid(row = row_ini+i, column = 0, sticky = NSEW)
#        
##        caseInpt = NameInpt.append("Input_" + str(i))
        caseInp = Entry(top, width = Input_width, font=(Font, Font_size))
        caseInp.grid(row = row_ini+i, column = 1, sticky=NSEW)
        if SPlist[i - (iniSP)] == 'Jacket part number':
            caseInp.insert(10, Input_PN.get())
        elif SPlist[i - (iniSP)] == 'Jacket customer part number':
            JCPN = Entry(top, width = Input_width, font=(Font, Font_size))
            JCPN.grid(row = row_ini+i, column = 1, sticky=NSEW)
            JCPN.insert(10, '0190-XXXXX')
            caseInp.insert(10, '0190-XXXXX')
        elif SPlist[i - (iniSP)] == 'Date of testing':
            caseInp.insert(10, Input_TD.get())
        elif SPlist[i - (iniSP)] == 'Run time (h:mm:ss)':
            i_min = 'B' + str(row_min + 1)
            i_max = 'B' + str(row_max)
            eq = str("=TEXT('Data log'!" + i_max + 
                          "-'Data log'!" + i_min + 
                          ', "h:mm:ss")')
            print(eq)
            caseInp.insert(20,  eq)
        else:
            caseInp.insert(10, str(i))
        
        cols.append(caseInp)
#        
    rows.append(cols)    
    
    Button(top, text = "Enter", font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=readCase).grid(row = iniSP+numSP, column= 0)
    
    Button(top, text = "Exit", font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "red", bg = "black", 
           command=top.destroy).grid(row = iniSP+numSP, column= 1)


#===============================================
def plot_2D(fcase, x, y, xlbl, ylbl):
    
#    x = np.linspace(x_min,x_max,points)
#    T = T_inf + Q/P * ( 1 - np.exp(-P*x) ) + ( T_i - T_inf ) * np.exp(-P*x)
    logging.debug(fcase)
    logging.debug(x)
    logging.debug(float(max(y)))
    plt.figure(fcase)
    
    plt.xticks( range(len(x)), x )
    
    plt.yticks(np.arange(0, 250, step=20))
    plt.ylim(0, 250)
#    ax.yaxis.set_minor_locator(minorLocator)
#    plt.scatter(range(len(y)), y )
    plt.plot(y,'ro')
    plt.xlabel(xlbl, fontsize=16)
    plt.ylabel(ylbl, fontsize=16)
    plt.xticks(color='k', size=16)
    plt.yticks(color='k', size=16)
#    plt.xlim(0, np.amax(x) )
#    
    plt.tight_layout()
    plt.show()

#===============================================    
def GenMatlabPlot():
    ULind = str( Input_TCF.get() )  #get the upper left index for plot
    LRind = str( Input_TCL.get() )  #get the lower right index for plot
    ULxy = coordinate_from_string( ULind )
    LRxy = coordinate_from_string( LRind )
    
    row_min = ULxy[1]
    row_max = LRxy[1]
    logging.debug( row_min )
    logging.debug( row_max )
    col_min = column_index_from_string( ULxy[0] )
    col_max = column_index_from_string( LRxy[0] )
    logging.debug( col_min )
    logging.debug( col_max )
    rows = {}
    
    
    total_col = ( col_max - col_min ) + 1 
    total_row = ( row_max - row_min ) + 1
    try:
        for j in range(col_min-1, col_max + 1):
            x = []
            y = []
            for i in range(row_min-1, row_max + 1):
                index = str(get_column_letter(j)) + str(i)
    #            logging.debug(index)
                logging.debug(ws[index].value)
                rows[i, j] = ws[index].value
                logging.debug( rows[i, j] )
                xtemp = str(rows[i, col_min-1])
                x.append(xtemp)
                if j != col_min-1:
                    if i == row_min-1:
                        ytemp = str(rows[i,j])
                        y.append(ytemp)
                    elif i > row_min-1:
                        ytemp = float(rows[i,j])
                        y.append(ytemp)
                    
            if j != col_min-1:     
                logging.debug(x[1:total_row+1])
                logging.debug(y[1:total_row+1])
                fcase = y[0]
                xlbl = 'T/C'
                ylbl = 'Temperature (°C)'
                plot_2D(fcase, x[1:total_row+1], y[1:total_row+1], xlbl, ylbl)  

    except KeyError:
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
            
#===============================================
def genLabTranPlot():
    
    global wbDL, row_min, row_max
    
    wbDL = wb[tablist[5]]
    print(wbDL.title)
    
    row_min = wbDL.min_row
    row_max = wbDL.max_row
    print('genLabTranPlot()', row_min, row_max)
    
    #estimation of the minimum and maximum of the column in data log tab
    
    col_min = wbDL.min_column
    col_max = wbDL.max_column
    print(get_column_letter(col_min), get_column_letter(col_max))
    

    total_cols = numTC + 3  
    try:
        for j in range(col_min + 1, col_min + 2 + total_cols):
#            print(j)
            for i in range(row_min, row_max + 1):
                index = str(get_column_letter(j)) + str(i)
            print(i, j)

        #=============================================================================
        #Charts for "Result" and "Chart" tabs#       
        data = Reference(wbDL, min_col=col_min + 2,              min_row=row_min, 
                               max_col=col_min + 2 + total_cols, max_row=row_max)
        chart = LineChart()
#        
        chart.title = "Data output of " + Input_PN.get() + "/" + cus_jack_PN
        chart.style = 13
        chart.x_axis.title = 'Run time'
        chart.y_axis.title = 'Temperature (°C)'
#        
        chart.add_data(data, titles_from_data = True)
        RT = Reference(wbDL, min_col=col_min+1, min_row=row_min+1, max_row=row_max)
#        #Setting T/C numbers as the x-axis coordinates
        chart.set_categories(RT)
        chart.height = 10 # default is 7.5
        chart.width = 40 # default is 15
#        
        # Style the lines
        for k in range(0, total_cols):
            s = chart.series[k]
#            s.marker.symbol = "circle"
#            s.marker.graphicalProperties.line.solidFill = "FF0000"
            #Setting colors on lines of charts
            s.graphicalProperties.line.solidFill = COLOR_INDEX[k]
#            s.graphicalProperties.line.noFill = True
#            logging.debug(k)
#        
        #"Chart" tab in excel. This is where to save the transient temperature    
        wbChart = wb[tablist[3]]
#        print( wbChart.title )#       
        #Save the chart in "Chart" tab
        wbChart.add_chart(chart, "B2")
        

        chartRes = deepcopy(chart)
        #"Chart" tab in excel. This is where to save the transient temperature 
        wbResult = wb[tablist[0]]        
#        print( wbResult.title )
        #Save the chart in "Result" tab
        wbResult.add_chart(chartRes, "A10")
        
        #=============================================================================
#        (min_row_DC_plot, max_row_DC_plot, min_col_DC_plot, max_col_DC_plot)
        #Charts for "Duty" tabs
        Ddata = Reference(wbDL, min_col=max_col_DC_plot, min_row=row_min, 
                               max_col=max_col_DC_plot, max_row=row_max)
        Dchart = LineChart()
#        
        Dchart.title = "Duty cycle of " + Input_PN.get() + "/" + cus_jack_PN
        Dchart.style = 13
        Dchart.x_axis.title = 'Run time'
        Dchart.y_axis.title = 'Duty cycle (%)'
        
        Dchart.add_data(Ddata, titles_from_data = True)
        DRT = Reference(wbDL, min_col=col_min+1, min_row=row_min+1, max_row=row_max) # time
        Dchart.set_categories(DRT)
        Dchart.height = 30 # default is 7.5
        Dchart.width = 40 # default is 15
        Dchart.series[0].graphicalProperties.line.solidFill = COLOR_INDEX[2]
        
        wbDuty = wb[tablist[4]]
        wbDuty.add_chart(Dchart, "B2")
        
    except KeyError:
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()    
        
#===============================================
def genLabSetUp():
    
    wbSP = wb[tablist[2]]
    
    print(wbSP)
    
    j = 1
    i = 3
    index = str(get_column_letter(j)) + str(i)
    wbSP[index].value = 'Setup'
    wbSP[index].font = openpyxl.styles.Font(name='Calibri', size=18, bold=True)  
    
    j = 1
    i_ini = 4
    for i in range(i_ini, i_ini+numSP):
        index = str(get_column_letter(j)) + str(i)
        wbSP[index].value = SPlist[i-i_ini]
        wbSP[index].font = openpyxl.styles.Font(name='Calibri', size=12, bold=True)        
        # To set alignment of text inside cell and text wrapping
        wbSP[index].alignment = openpyxl.styles.Alignment(horizontal='right', 
                vertical='center', wrap_text=True)
        # Fill background of the cell
        wbSP[index].fill = PatternFill(fgColor="CCCCCC", fill_type = "solid") 
        
        wbSP[index].border = thin_border
            
        index_val = str(get_column_letter(j+1)) + str(i)
        wbSP[index_val].value = excelCaName[i - i_ini + numC]
        wbSP[index_val].border = thin_border
    
    # set the width of the column
    wbSP.column_dimensions[get_column_letter(j)].width = 50
    # set the width of the next column 
    wbSP.column_dimensions[get_column_letter(j+1)].width = 25
    
#===============================================
def genLabResult():
    
    global numTC    
    
    wbResult = wb[tablist[0]]
    print( wbResult.title )
    
    
    TCF_index = str( Input_TCF.get() )  #get the upper left index for plot
    print("1st TC number ", TCF_index)
    TCL_index = str( Input_TCL.get() )  #get the lower right index for plot
    print(TCL_index)
    
    #insert the image in the sheet
    img = openpyxl.drawing.image.Image('BriskHeat logo.jpg')
    wbResult.add_image(img, 'A1')
    
    inirow = 35
    inicol = 3
#    defaultCase = 2
    numTC = int(TCL_index) - int(TCF_index) + 1
    j = inicol
#    print(name)

    # Cell of Part number
    j = inicol
    for i in range(inirow - 2, inirow + numTC + 3):
        index = str(get_column_letter(j)) + str(i)
        
        if i == (inirow - 2):   # Cell of 'Part number'
            ind_rang = index + ":" + str(get_column_letter(j + numC)) + str(i) 
            print(ind_rang)
            wbResult.merge_cells(ind_rang, "")        
                        
        elif i == (inirow - 1): # Cell of 'Thermocouple (°C)'
            wbResult[index].value = ReTabList[1]
        elif i in range(inirow, inirow + numTC):
            wbResult[index].value = ReTabList[3] + ' ' + str( int(TCF_index) + (i - inirow) )      
        elif i == (inirow + numTC):     # Cell of 'Control'
            wbResult[index].value = ReTabList[5]
        elif i == (inirow + numTC + 1): # Cell of 'Duty cycle'
            wbResult[index].value = ReTabList[6]
#        elif i == (inirow + numTC + 2):  # Cell of 'Maximum wattage (W)'
#            wbResult[index].value = ReTabList[7]
#        elif i == (inirow + numTC + 3):  # Cell of 'Power (W)'
#            wbResult[index].value = ReTabList[8]
        elif i == (inirow + numTC + 2):   # Cell of 'Notes'
            wbResult[index].value = ReTabList[9]
            
        wbResult[index].font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
        
        # To set alignment of text inside cell and text wrapping
        wbResult[index].alignment = openpyxl.styles.Alignment(horizontal='center', 
                vertical='center', wrap_text=True)
        # Fill background of the cell
        wbResult[index].fill = PatternFill(fgColor="CCCCCC", fill_type = "solid")
        
        wbResult[index].border = thin_border

    # set the width of the column 
    wbResult.column_dimensions[get_column_letter(j)].width = 25
        
    # set the height of the row for note column
    wbResult.row_dimensions[(inirow + numTC + 2)].height = 70
    i = inirow - 2
    for j in range(inicol+1, inicol + 1 + numC ):
        index = str(get_column_letter(j)) + str(i)
        wbResult[index].border = thin_border
    
    j = inicol
    index = str(get_column_letter(j)) + str(i)
    wbResult[index].value = Input_PN.get()

    i = inirow - 1
    for j in range(inicol+1, inicol + 1 + numC ):
#        print(j)
        index = str(get_column_letter(j)) + str(i)
        wbResult[index].value = excelCaName[j- inicol - 1]
        
        wbResult[index].font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)        
        # To set alignment of text inside cell and text wrapping
        wbResult[index].alignment = openpyxl.styles.Alignment(horizontal='center', 
                vertical='center', wrap_text=True)
        # Fill background of the cell
        wbResult[index].fill = PatternFill(fgColor="CCCCCC", fill_type = "solid")  
        
        wbResult[index].border = thin_border
    
    for j in range(inicol+1, inicol + 1 + numC ):
        for i in range(inirow, inirow + numTC + 3):
            index = str(get_column_letter(j)) + str(i)
            wbResult[index].border = thin_border
    
#===============================================        
def GenLabTest():
    
    global wb
    #openpyxl.load_workbook() function takes in the filename and returns value of the workbook data type.
    wb = openpyxl.load_workbook(path_UTR)  
    
    OpenExcel_UTR(path_UTR)
    OpenExcel_RAW(path_RAW)
    OpenExcel_DC(path_DC)
#    Reading_PDF(path_PDF)
    
    genLabResult()
    genLabTranPlot()
    genLabSetUp()    
    
    file_name = 'Testing Result for ' + Input_PN.get() + '-' + Input_TNum.get() + '-' + Input_TD.get() + '.xlsx'
    logging.debug('Generating the file successful')
    wb.save(file_name)    
    
    top = Toplevel()
    w = 600
    h = 100 
    top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
    top.title('Done')

    msg = Message(top, text='Successfully generating the test template in ' + file_name + '!', width=600)
    msg.pack()
    
    # Play Windows exit sound.
    winsound.PlaySound("SystemExit", winsound.SND_ALIAS)

#    button = Button(top, text="Exit", command=top.destroy)
    button = Button(top, text="Exit GUI", command=des)
    button.pack() 
    
    return None
    
#===============================================    
def OpenExcel_UTR(path_UTR):
       
    for i in wb:
        if i.title != 'UTR':
            wb.remove(i)
#            logging.debug( i )
    
    for i in range(len(tablist)):
        if tablist[i] != 'UTR':
            wb.create_sheet(tablist[i])
        
    # Setting the background color
    for j in range(col_min, col_max + 1):
            for i in range(row_min, row_max + 1):
                index = str(get_column_letter(j)) + str(i)   
                wb['Result'][index].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", 
                      fill_type = "solid")
                wb['Setup'][index].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                      fill_type = "solid")
                    
                wb['Chart'][index].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC",
                      fill_type = "solid")
                wb['Duty'][index].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC",
                      fill_type = "solid")
                wb['Picture'][index].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC"
                      , fill_type = "solid")
    
#===============================================
def OpenExcel_RAW(path_RAW):
    
    global max_col_TC, row_max

    offsetRow = 4
    print('=====================================')
    print( path_RAW )
    print( wb.active )
    wbDL = wb[tablist[5]]
    print( wbDL.title )
    print('=====================================')
    
    #Reaing CSV files
    f = open(path_RAW, 'r')
    
    with f:
        
        reader = csv.reader(f)
        print(reader)
        i = 1
        for row in reader:  #reading row by row
            
            if i < 2:
                j = 1   #reset the number of column
                for e in row:   #reading column by column in a single row
                    index = str(get_column_letter(j)) + str(i)
#                    print(index, e)
#                print(i, j)
                    #Convert the data form "text" type to "numeric" type
                    wbDL[index].value = pd.to_numeric(e, downcast='float', errors='ignore')
                    j += 1
                    
            elif i > offsetRow:
                j = 1   #reset the number of column
                for e in row:   #reading column by column in a single row
                    index = str(get_column_letter(j)) + str(i - offsetRow + 1)
                    #Convert the data form "text" type to "numeric" type
                    wbDL[index].value = pd.to_numeric(e, downcast='float', errors='ignore')
#                    print(index, e)
                    j += 1
                
            i += 1
    
    print("The maximum column number of T/C is ", get_column_letter(j-1) )
    max_col_TC = j - 1
    max_row_TC_raw = i - 1
    print("The maximum row number of T/C is ", max_row_TC_raw )
    row_max = max_row_TC_raw
    
            
#===============================================
def OpenExcel_DC(path_DC):
    
    global min_row_DC_plot, max_row_DC_plot, min_col_DC_plot, max_col_DC_plot
    
    offsetRow = 2
    
    print( path_DC )
    #openpyxl.load_workbook() function takes in the filename and returns value of the workbook data type.
    wbDC = openpyxl.load_workbook(path_DC)  
    print(wbDC.active.title)
    wbDCsheet = wbDC[wbDC.active.title]
    print(wbDCsheet.max_row)
    
#    call the data log sheet
    wbDL = wb[tablist[5]]
    
    print("zone", Input_DC.get())
    
    num_zone = int( Input_DC.get() )
    name_zone = ['Control', 'Setpoint', 'Duty cycle (%)']
#    cols_per_zone = [str(4 * num_zone-1), str(4 * num_zone), str(4 * num_zone + 1)]
    cols_per_zone = [(4 * num_zone-1), (4 * num_zone), (4 * num_zone + 1)]
    
    print(cols_per_zone)
    
    col_index_data_log_ini = (max_col_TC + 1)
    print(col_index_data_log_ini)
    
    
    #Reading the data from one excel and write them into the final data log sheet in the result template
    max_row_DC = wbDCsheet.max_row
    i = 1
    for row in range(max_row_DC):
        
        if i < offsetRow:
            for j in range (0, 3):
                index_result = str( get_column_letter(col_index_data_log_ini+j) ) + str(i)
                index_DC = str(get_column_letter( cols_per_zone[j] )) + str(i)
#                wbDL[index_result].value = wbDCsheet[index_DC].value
                wbDL[index_result].value = name_zone[j]
#                print(wbDL[index_result].value, wbDCsheet[index_DC].value)
                
        elif i > offsetRow:
            for j in range (0, 3):
                index_result = str( get_column_letter(col_index_data_log_ini+j) ) + str(i - 1) # - 1 is to take a row from DC raw data
                index_DC = str(get_column_letter( cols_per_zone[j] )) + str(i)
                wbDL[index_result].value = wbDCsheet[index_DC].value
#                print(index_result, wbDCsheet[index_DC].value)

        i += 1

    min_row_DC_plot = 2
    max_row_DC_plot = max_row_DC - 1
    min_col_DC_plot = col_index_data_log_ini
    max_col_DC_plot = col_index_data_log_ini + 2
#    print(min_row_DC_plot, max_row_DC_plot, min_col_DC_plot, max_col_DC_plot)
    
#===============================================
def Reading_PDF(path_PDF):
    print(path_PDF)
    
    text = textract.process(path_PDF)
    print(text)
    
    
#===============================================
#This is where we lauch the file manager bar.
def OpenFileUTR():
    
    global path_UTR
#    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All Excel Files", "*.xl*;*.xlsx*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )                 
    # Make sure clear entry whenever path is re-selected.
    UTR_Path = Entry(root, width = '100', font=(Font, Font_size))
    UTR_Path.grid(row = 0, column = 1, sticky=NSEW)
    UTR_Path.insert(80, ' ')
    
    path_UTR = root.filename
    logging.debug(root.filename)
    logging.debug(path_UTR)  
    logging.debug(ws)

    UTR_Path.insert(50, root.filename)
    logging.debug(UTR_Path.get())
    
    if len(UTR_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
    elif len(UTR_Path.get()) > 1:
        print('=====================================')
        print(path_UTR)
        print('=====================================')
#        logging.debug('=====================================')
#        logging.debug(path_UTR)
#        logging.debug('=====================================')
        
#    logging.debug(ws)
#===============================================
def OpenFileRAW():
    
    global path_RAW
#    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All Excel Files", "*.csv*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )
 
    # Make sure clear entry whenever path is re-selected.
    RAW_Path = Entry(root, width = '100', font=(Font, Font_size))
    RAW_Path.grid(row = 1, column = 1, sticky=NSEW)
    RAW_Path.insert(80, ' ')
                           
    path_RAW = root.filename
    logging.debug(root.filename)
    logging.debug(path_RAW)
    
#    logging.debug(ws)
    RAW_Path.insert(50, root.filename)
    logging.debug(RAW_Path.get())
    
    if len(RAW_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
    elif len(RAW_Path.get()) > 1:
        print('=====================================')
        print(path_RAW)
        print('=====================================')
        OpenExcel_RAW(path_RAW) #To read the max row number of raw TC data here.
        
#        logging.debug('=====================================')
#        logging.debug(path_RAW)
#        logging.debug('=====================================')
        
    

#===============================================
def OpenFileDC():
    global path_DC
    #    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All Excel Files", "*.xl*;*.xlsx*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )
                           
    
    # Make sure clear entry whenever path is re-selected.
    DC_Path = Entry(root, width = '100', font=(Font, Font_size))
    DC_Path.grid(row = 2, column = 1, sticky=NSEW)
    DC_Path.insert(80, ' ')
    
    path_DC = root.filename
    logging.debug(path_DC)
    DC_Path.insert(50, root.filename)
    
    if len(DC_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
    elif len(DC_Path.get()) > 1:
        print('=====================================')
        print(path_DC)
        print('=====================================')
    
    
#===============================================    
def OpenFilePDF():
    
    global path_PDF
#    The askopenfilename function to creates an file dialog object.
    root.filename = askopenfilename(initialdir=cwd,
                           filetypes =(("All PDF Files", "*.pdf*")
                           ,("All Files","*.*")),
                           title = "Choose a file."
                           )
                           
    path_PDF = root.filename
    PDF_Path.insert(50, path_PDF)
    
    if len(PDF_Path.get()) <= 1:
        logging.debug('Please select a file')
        
        top = Toplevel()
        w = 150
        h = 100   
        top.geometry("%dx%d+%d+%d" % (w, h, wids/2, heis/2))
        top.title('Warning')

        msg = Message(top, text='Please select a file')
        msg.pack()

        button = Button(top, text="Exit", command=top.destroy)
        button.pack()
        
    elif len(PDF_Path.get()) > 1:
        print('=====================================')
        print(path_PDF)
        print('=====================================')
    

#===============================================
def img_label(imgPath, r, col):
#    print(imgPath)
    photo = PhotoImage(file = imgPath)
    label = Label(root, image=photo, bg = 'lightgrey')
    label.image = photo
    label.grid(row = r, column = col, columnspan=5, rowspan=2,sticky=W+N+S, padx=5, pady=5)
    
#===============================================
def clear_all():
#    gui.grid_forget()

    for widget in root.winfo_children():      # get all children of root
        if widget.winfo_class() == 'Entry':   # if the class is Entry
            widget.delete(0,END)              # reset its value
        elif widget.winfo_class() == 'Label': # get Label widgets
            widget.config(bg="light grey")         # and change them as well
            
    return None
#===============================================            
def des():
    logging.debug('End of program')
    root.destroy()

#===============================================
#def trigger_1_2():
#    
#    readCase()
#    
#    top.destroy
  
#===============================================            
if __name__ == "__main__":
    global wids, heis, rx, ry, w, h
    ws = {}
    
    root = Tk() #create one root widget for this program
    # get screen width and height
    wids = root.winfo_screenwidth() # width of the screen
    heis = root.winfo_screenheight() # height of the screen
    rx = root.winfo_x()
    ry = root.winfo_y()
    logging.debug( wids )
    logging.debug( heis )

#    =========================================
    Title = root.title( "BH lab test template generator (Technical support: ext. 1235)")
    root.configure(background="Dark grey")
    root.geometry("1000x512")
    
    #This button serves as a label to select a file
#    Label(root, text = 'File', font=(Font, Font_size), 
#          height = Label_height, width = Label_width, bd=Label_bd, 
#          relief="groove").grid(row = 0, column = 0, sticky = W)
#    ========================================= UTR form
    UTR_Path = Entry(root, width = '100', font=(Font, Font_size))
    UTR_Path.grid(row = 0, column = 1, sticky=NSEW)
    UTR_Path.insert(80, ' ')
    
    tk.Button(root, text='Selecting a UTR from (.xlsx)', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFileUTR
              ).grid(row = 0 , column = 0, sticky = NSEW)
#    ========================================= RAW form
    
    RAW_Path = Entry(root, width = '100', font=(Font, Font_size))
    RAW_Path.grid(row = 1, column = 1, sticky=NSEW)
    RAW_Path.insert(80, ' ')
    
    tk.Button(root, text='Selecting a data-T/C (.csv)', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFileRAW
              ).grid(row = 1 , column = 0, sticky = NSEW)
    
#    ========================================= Duty cycle (DC) form
    
    DC_Path = Entry(root, width = '100', font=(Font, Font_size))
    DC_Path.grid(row = 2, column = 1, sticky=NSEW)
    DC_Path.insert(80, ' ')
    
    tk.Button(root, text='Selecting a data-duty cycle (.xlsx)', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
              relief="groove", fg = "black", bg = "light grey", command=OpenFileDC
              ).grid(row = 2 , column = 0, sticky = NSEW)
#    =========================================  
#    PDF_Path = Entry(root, width = '100', font=(Font, Font_size))
#    PDF_Path.grid(row = 2, column = 1, sticky=NSEW)
#    PDF_Path.insert(80, ' ')
#    
#    tk.Button(root, text='Selecting a PDF data (.pdf)', font=(Font, Font_size), bd=Label_bd, width = Label_width, 
#              relief="groove", fg = "black", bg = "light grey", command=OpenFilePDF
#              ).grid(row = 2 , column = 0, sticky = NSEW)  
    
#    Enter zone of duty cycle
    Label(root, text = 'Zone #', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 3, column = 0, sticky = NSEW)
    
    Input_DC = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_DC.grid(row = 3, column = 1, sticky=W)
    Input_DC.insert(10, "1")     
    
#    Enter part number of jacket
    Label(root, text = 'Jacket part number', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 4, column = 0, sticky = NSEW)
    
    Input_PN = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_PN.grid(row = 4, column = 1, sticky=W)
    Input_PN.insert(10, "UAPMXX0XXXSN-XXX")    

#   Enter a number of the test
    Label(root, text = 'Test #', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 5, column = 0, sticky = NSEW)
    
    Input_TNum = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_TNum.grid(row = 5, column = 1, sticky=W)
    Input_TNum.insert(10, "1")  
#        
    #Enter the test date
    Label(root, text = 'Date of testing (MMDDYYYY)', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 6, column = 0, sticky = NSEW)
    
    Input_TD = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_TD.grid(row = 6, column = 1, sticky=W)
    Input_TD.insert(10, curDate)    
    
#    #Enter the # of the first T/C
    Label(root, text = 'T/C # (First)', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 7, column = 0, sticky = NSEW)
    
    Input_TCF = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_TCF.grid(row = 7, column = 1, sticky=W)
    Input_TCF.insert(10, '1')
#    
#    #Enter the # of the last T/C
    Label(root, text = 'T/C # (Last)', font=(Font, Font_size), 
          height = Label_height, width = Label_width, bd=Label_bd, 
          relief="groove").grid(row = 8, column = 0, sticky=NSEW)
    
    Input_TCL = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_TCL.grid(row = 8, column = 1, sticky=W)
    Input_TCL.insert(10, '6')
    
#     Selecting case
    Button(root, text = 'Entry of test cases', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "black", bg = "light grey", 
           command=popCase).grid(row = 9, column= 0, sticky=NSEW)
    Input_CT = Entry(root, width = Input_width, font=(Font, Font_size))
    Input_CT.grid(row = 9, column = 1, sticky=N+S+W)
    Input_CT.insert(10, '2')
    

#    
    # generate the button for creaing input table (Now)
#    Button(root, text = 'Generating plots in an excel file', font=(Font, Font_size), bd=Label_bd, relief="groove", 
#           fg = "white", bg = "dark green", 
#           command=GenPlot).grid(row = 5, column= 0, sticky=NSEW)
    
    Button(root, text = 'Generating a lab test template', font=(Font, Font_size), bd=Label_bd, relief="groove", 
           fg = "white", bg = "dark green", 
           command=GenLabTest).grid(row = 10, column= 0, sticky=NSEW)

    Buttn_refresh = tk.Button(root, text='Clear all entries', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "Black", bg = 
                       "Red", command=clear_all).grid(row = 11, column = 0, sticky=NSEW)
    
    Buttn_end = tk.Button(root, text='End/Exit', font=(Font, Font_size), bd=Label_bd, 
                       width = Label_width, relief="groove", fg = "black", bg = 
                       "red", command=des).grid(row = 12, column = 0, sticky=NSEW)
    
#    imgPath = "BH_logo.gif"
#    img_label(imgPath, 12, 0)
#
    root.mainloop()
    logging.debug('End of program')
